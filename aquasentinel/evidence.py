"""Schema-driven, read-only evidence analysis for AquaSentinel."""
from __future__ import annotations

import csv
import hashlib
import json
import math
import re
from dataclasses import dataclass, field
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

import numpy as np
from sklearn.ensemble import IsolationForest
from sklearn.preprocessing import StandardScaler

SUPPORTED_SUFFIXES = {".csv", ".json", ".jsonl", ".xlsx"}
DOMAIN_HINTS = {
    "SCADA / OT": {
        "scada",
        "plc",
        "rtu",
        "hmi",
        "protocol",
        "network",
        "command",
        "valve",
        "pump",
        "setpoint",
        "control",
    },
    "WATER QUALITY": {
        "ph",
        "turbidity",
        "chlorine",
        "conductivity",
        "tds",
        "salinity",
        "quality",
        "contamination",
        "permeate",
    },
    "DESALINATION / PROCESS": {
        "membrane",
        "ro",
        "pressure",
        "flow",
        "recovery",
        "fouling",
        "feed",
        "brine",
        "temperature",
        "train",
    },
    "ENERGY / RESOURCE": {
        "energy",
        "power",
        "kwh",
        "efficiency",
        "demand",
        "consumption",
        "resource",
        "recovery",
    },
    "MAINTENANCE / ASSET": {
        "maintenance",
        "asset",
        "health",
        "condition",
        "failure",
        "workorder",
        "inspection",
        "service",
    },
    "ACCESS / IDENTITY": {
        "user",
        "identity",
        "badge",
        "access",
        "login",
        "role",
        "failed",
        "mfa",
        "operator",
    },
    "COMPLIANCE / AUDIT": {
        "audit",
        "compliance",
        "approval",
        "change",
        "policy",
        "control",
        "evidence",
        "review",
        "record",
    },
}
TIME_HINTS = {
    "timestamp",
    "time",
    "datetime",
    "date",
    "eventtime",
    "observed",
    "created",
    "recorded",
}


def _tokens(value: str) -> set[str]:
    return {part for part in re.split(r"[^a-z0-9]+", value.lower()) if part}


def _as_float(value: Any) -> float | None:
    if value is None or isinstance(value, bool):
        return None
    try:
        text = str(value).strip().replace(",", "")
        if not text:
            return None
        number = float(text)
        return number if math.isfinite(number) else None
    except (TypeError, ValueError):
        return None


def _parse_time(value: Any) -> datetime | None:
    if value is None or str(value).strip() == "":
        return None
    if isinstance(value, datetime):
        parsed = value
    else:
        text = str(value).strip().replace("Z", "+00:00")
        try:
            parsed = datetime.fromisoformat(text)
        except ValueError:
            return None
    if parsed.tzinfo is None:
        parsed = parsed.replace(tzinfo=timezone.utc)
    return parsed.astimezone(timezone.utc)


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


@dataclass
class EvidenceFile:
    path: str
    name: str
    file_type: str
    rows: int
    columns: list[str]
    domain: str
    confidence: float
    domain_scores: dict[str, float] = field(default_factory=dict)
    numeric_features: list[str] = field(default_factory=list)
    missing_cells: int = 0
    total_cells: int = 0
    anomaly_score: float = 0.0
    anomaly_flags: int = 0
    analyzed_rows: int = 0
    notes: list[str] = field(default_factory=list)
    sha256: str = ""
    timestamp_field: str | None = None
    time_start: str | None = None
    time_end: str | None = None
    analysis_method: str = "schema-only"

    @property
    def missing_pct(self) -> float:
        if not self.total_cells:
            return 0.0
        return 100.0 * self.missing_cells / self.total_cells


@dataclass
class EvidencePackage:
    datasets: list[EvidenceFile]
    risk_score: float
    risk_level: str
    correlations: list[str] = field(default_factory=list)

    @property
    def total_rows(self) -> int:
        return sum(item.rows for item in self.datasets)

    @property
    def total_flags(self) -> int:
        return sum(item.anomaly_flags for item in self.datasets)


def _read_rows(path: Path) -> list[dict[str, Any]]:
    suffix = path.suffix.lower()
    if suffix == ".csv":
        with path.open("r", encoding="utf-8-sig", newline="") as handle:
            return [dict(row) for row in csv.DictReader(handle)]

    if suffix == ".json":
        data = json.loads(path.read_text(encoding="utf-8"))
        if isinstance(data, list):
            return [dict(row) for row in data if isinstance(row, dict)]
        if isinstance(data, dict):
            for value in data.values():
                if isinstance(value, list) and all(
                    isinstance(row, dict) for row in value
                ):
                    return [dict(row) for row in value]
            return [data]

    if suffix == ".jsonl":
        rows: list[dict[str, Any]] = []
        for line in path.read_text(encoding="utf-8").splitlines():
            if not line.strip():
                continue
            value = json.loads(line)
            if isinstance(value, dict):
                rows.append(value)
        return rows

    if suffix == ".xlsx":
        from openpyxl import load_workbook

        workbook = load_workbook(path, read_only=True, data_only=True)
        sheet = workbook.active
        values = sheet.iter_rows(values_only=True)
        header = [
            str(value).strip() if value is not None else ""
            for value in next(values, ())
        ]
        return [
            {
                header[index]: value
                for index, value in enumerate(row)
                if index < len(header) and header[index]
            }
            for row in values
        ]

    raise ValueError(f"Unsupported evidence type: {suffix}")


def _infer_domain(columns: list[str]) -> tuple[str, float, dict[str, float]]:
    tokens = set().union(*(_tokens(column) for column in columns)) if columns else set()
    scores: dict[str, float] = {}
    for domain, hints in DOMAIN_HINTS.items():
        matches = len(tokens & hints)
        scores[domain] = min(1.0, matches / 3.0) if matches else 0.0

    domain = max(scores, key=scores.get) if scores else "GENERAL / UNKNOWN"
    confidence = scores.get(domain, 0.0)
    if not confidence:
        return "GENERAL / UNKNOWN", 0.0, scores
    return domain, confidence, scores


def _infer_timestamp(
    rows: list[dict[str, Any]],
    columns: list[str],
) -> tuple[str | None, str | None, str | None]:
    best: str | None = None
    best_values: list[datetime] = []
    best_ratio = 0.0

    for column in columns:
        hint = bool(_tokens(column) & TIME_HINTS)
        parsed = [_parse_time(row.get(column)) for row in rows]
        values = [value for value in parsed if value is not None]
        ratio = len(values) / len(rows) if rows else 0.0
        if values and ratio >= 0.6 and (hint or ratio > best_ratio):
            best = column
            best_values = values
            best_ratio = ratio

    if not best:
        return None, None, None
    return (
        best,
        min(best_values).isoformat(),
        max(best_values).isoformat(),
    )


def _discover_numeric(
    rows: list[dict[str, Any]],
    columns: list[str],
) -> list[str]:
    usable: list[str] = []
    for column in columns:
        present = [
            value
            for value in (_as_float(row.get(column)) for row in rows)
            if value is not None
        ]
        if len(present) < max(5, int(len(rows) * 0.6)) or len(set(present)) < 2:
            continue
        column_tokens = _tokens(column)
        if column_tokens & {"id", "identifier", "timestamp", "time", "date"}:
            continue
        near_unique = len(set(present)) / max(1, len(present)) > 0.98
        identifier_hint = any(
            token in column_tokens
            for token in {"number", "code", "sequence", "index"}
        )
        if near_unique and identifier_hint:
            continue
        usable.append(column)
    return usable


def _robust_anomaly(matrix: list[list[float]]) -> tuple[float, int]:
    array = np.asarray(matrix, dtype=float)
    median = np.median(array, axis=0)
    mad = np.median(np.abs(array - median), axis=0)
    mad = np.where(mad < 1e-9, 1.0, mad)
    row_scores = np.max(0.6745 * np.abs(array - median) / mad, axis=1)
    flags = int(np.sum(row_scores > 3.5))
    pressure = min(100.0, float(np.percentile(row_scores, 90)) / 7.0 * 100.0)
    return round(pressure, 1), flags


def analyze_file(path: str | Path) -> EvidenceFile:
    evidence_path = Path(path).expanduser().resolve()
    if not evidence_path.is_file():
        raise FileNotFoundError(evidence_path)
    if evidence_path.suffix.lower() not in SUPPORTED_SUFFIXES:
        raise ValueError(f"Unsupported evidence type: {evidence_path.suffix}")

    rows = _read_rows(evidence_path)
    columns = list(dict.fromkeys(key for row in rows for key in row.keys()))
    domain, confidence, scores = _infer_domain(columns)
    numeric = _discover_numeric(rows, columns)
    total = len(rows) * len(columns)
    missing = sum(
        1
        for row in rows
        for column in columns
        if row.get(column) in (None, "")
    )
    timestamp_field, time_start, time_end = _infer_timestamp(rows, columns)

    result = EvidenceFile(
        path=str(evidence_path),
        name=evidence_path.name,
        file_type=evidence_path.suffix.lower(),
        rows=len(rows),
        columns=columns,
        domain=domain,
        confidence=confidence,
        domain_scores=scores,
        numeric_features=numeric,
        missing_cells=missing,
        total_cells=total,
        sha256=_sha256(evidence_path),
        timestamp_field=timestamp_field,
        time_start=time_start,
        time_end=time_end,
    )

    if not rows:
        result.notes.append("No records were available for analysis.")
        return result
    if not numeric:
        result.notes.append(
            "No suitable numeric features were discovered; schema profiling only."
        )
        return result

    matrix: list[list[float]] = []
    for row in rows:
        vector = [_as_float(row.get(column)) for column in numeric]
        if all(value is not None for value in vector):
            matrix.append([float(value) for value in vector if value is not None])

    result.analyzed_rows = len(matrix)
    if len(matrix) < 5:
        result.notes.append(
            "Too few complete numeric records for a defensible anomaly finding."
        )
        return result

    if len(matrix) < 20:
        result.anomaly_score, result.anomaly_flags = _robust_anomaly(matrix)
        result.analysis_method = "robust-MAD"
    else:
        scaled = StandardScaler().fit_transform(matrix)
        model = IsolationForest(
            n_estimators=160,
            contamination="auto",
            random_state=133,
        )
        predictions = model.fit_predict(scaled)
        raw = -model.score_samples(scaled)
        low = float(min(raw))
        high = float(max(raw))
        normalized = sorted(
            100.0 * (float(value) - low) / (high - low) if high > low else 0.0
            for value in raw
        )
        index = max(
            0,
            min(
                len(normalized) - 1,
                math.ceil(0.90 * len(normalized)) - 1,
            ),
        )
        result.anomaly_score = round(normalized[index], 1)
        result.anomaly_flags = sum(1 for value in predictions if value == -1)
        result.analysis_method = "IsolationForest"

    result.notes.append(
        "Anomaly indicators are not proof of contamination, compromise, or unsafe water."
    )
    return result


def _time_correlations(datasets: list[EvidenceFile]) -> list[str]:
    findings: list[str] = []
    timed = [dataset for dataset in datasets if dataset.time_start and dataset.time_end]
    for index, left in enumerate(timed):
        for right in timed[index + 1 :]:
            left_start = datetime.fromisoformat(left.time_start)
            left_end = datetime.fromisoformat(left.time_end)
            right_start = datetime.fromisoformat(right.time_start)
            right_end = datetime.fromisoformat(right.time_end)
            overlap_start = max(left_start, right_start)
            overlap_end = min(left_end, right_end)
            if overlap_start <= overlap_end:
                findings.append(
                    f"{left.name} ↔ {right.name}: overlapping UTC evidence window "
                    f"{overlap_start.isoformat()} to {overlap_end.isoformat()}"
                )
    return findings


def analyze_package(paths: list[str | Path]) -> EvidencePackage:
    datasets = [analyze_file(path) for path in paths]
    weighted = [
        (dataset.anomaly_score, dataset.analyzed_rows or dataset.rows)
        for dataset in datasets
        if dataset.rows
    ]
    denominator = sum(weight for _, weight in weighted)
    risk = (
        round(
            sum(score * weight for score, weight in weighted) / denominator,
            1,
        )
        if denominator
        else 0.0
    )
    if risk < 40:
        level = "NORMAL"
    elif risk < 70:
        level = "ELEVATED"
    else:
        level = "HIGH"
    return EvidencePackage(
        datasets=datasets,
        risk_score=risk,
        risk_level=level,
        correlations=_time_correlations(datasets),
    )
